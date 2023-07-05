import os
import xlrd
from openpyxl import Workbook
from openpyxl.utils import datetime as openpyxl_datetime
from datetime import datetime as python_datetime
from tkinter import simpledialog

dir_with_files = simpledialog.askstring(title = 'Folder path prompt', prompt = "Please type in the full path of the folder containing your files:    ")

for folder in os.listdir(dir_with_files):
    folder = os.path.join(dir_with_files, folder)
    try:
        os.chdir(folder)
    except:
        continue

    for file in os.listdir(folder):

        # Open the .xls file for reading
        xls_path = file
        try:
            workbook = xlrd.open_workbook(xls_path)
        except:
            continue
        # Create a new workbook for writing
        new_workbook = Workbook()


        # Iterate over sheets in the original workbook
        for sheet in workbook.sheets():
            # Create a new sheet in the new workbook
            new_sheet = new_workbook.create_sheet(title=sheet.name)

            # Iterate over rows and columns in the original sheet
            for row in range(sheet.nrows):
                for col in range(sheet.ncols):
                    cell_value = sheet.cell_value(row, col)

                    # Check if the cell contains a date or time value
                    if sheet.cell_type(row, col) == xlrd.XL_CELL_DATE:
                        try: # In case there is an overflow error
                            # Convert Excel date/time value to Python datetime object
                            cell_value = xlrd.xldate_as_datetime(cell_value, workbook.datemode)
                        except: 
                            pass

                    # Write the cell value to the new sheet
                    new_sheet.cell(row=row+1, column=col+1).value = cell_value

                    # Adjust the cell format if it contains a date or time value
                    if isinstance(cell_value, python_datetime):
                        new_sheet.cell(row=row+1, column=col+1).number_format = 'yyyy-mm-dd hh:mm:ss'

        # Save the new workbook as an .xlsx file
        xlsx_path = file + '.xlsx'
        new_workbook.save(xlsx_path)
        os.remove(file)
