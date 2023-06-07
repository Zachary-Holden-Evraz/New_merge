# RPT Merge
# Note that this was created in Python and made into a windows executable file

import os, sys
from sys import exit
import fnmatch
import shutil
import pandas as pd
from threading import Thread
import openpyxl as pyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import *
from tkinter import simpledialog
from tkinter.filedialog import askopenfilename
import re
from datetime import datetime, time

# Ask the User to give the files - this will allow this to work on any Windows computer
main_file = askopenfilename(title = 'File you want to update') # Excel File we want to add all the data into
dir_with_files = simpledialog.askstring(title = 'Folder path prompt', prompt = "Please type in the full path of the folder containing your files:    ")

def Merge_RPT_files():
    global main_file_check; global main_data; global main_sheet; global workbook; global worksheet; global row_count; global column_count
    extra_save = 0 # For saving a backup automatically 

    # Setting up the main file everything will go into
    if main_file_check == 0:
        main_data = pyxl.load_workbook(main_file)
        main_sheet = main_data.worksheets[0]

        # Setting up the active workbook
        global workbook
        workbook = Workbook()
        worksheet = workbook.active
        row_count = main_sheet.max_row
        column_count = main_sheet.max_column

    if stop == 0:
        stop_program()

    # Make sure data in the main file stays (because this will overwite existing file)
    for i in range (1, row_count + 1):
        for j in range (1, column_count + 1):
            # reading cell value from source file
            c = main_sheet.cell(row = i, column = j)
            # writing the read value to destination file
            worksheet.cell(row = i, column = j).value = c.value
            # worksheet.cell(row = i, column = j).fill  = c.fill # keep any background filling
            main_file_check = 1

    if stop == 0:
        stop_program()

    for folder in os.listdir(dir_with_files):
        folder = os.path.join(dir_with_files, folder)
        try:
            os.chdir(folder)
        except:
            continue
        for file in os.listdir(folder):
            try:
                book = pyxl.load_workbook(file, data_only = True)
            except:
                continue
            # names = book.sheetnames ### Does not work since there are charts in some workbooks (Counts as a sheetname but not a worksheet)
            names = [sheet.title for sheet in book.worksheets]
            # Have to get index for specific sheet
            i = 0
            sheets = []
            for name in names:
                if name == 'INSP RPT 1':
                    sheets.append(i)
                elif name == 'INSP RPT 1 (2)':
                    sheets.append(i)
                elif name == 'INSP RPT 2':
                    sheets.append(i)
                elif name == 'INSP RPT 2 (2)':
                    sheets.append(i)
                elif name == 'INSP RPT 1 (8hr)':
                    sheets.append(i)
                elif name == 'INSP RPT 1 (8hr) (2)':
                    sheets.append(i)
                elif name == 'INSP RPT 1 (12hr)':
                    sheets.append(i)
                elif name == 'INSP RPT 1 (12hr) (2)':
                    sheets.append(i)
                elif name == 'INSP RPT 2 (8hr)':
                    sheets.append(i)
                elif name == 'INSP RPT 2 (8hr) (2)':
                    sheets.append(i)
                elif name == 'INSP RPT 2 (12hr)':
                    sheets.append(i)
                elif name == 'INSP RPT 2 (12hr) (2)':
                    sheets.append(i)
                i += 1
            if i >= len(names)-1:
                pass
            # sheet = book.worksheets[i] ##### Need to make the script work for all sheets in sheets

            # All sheets that need to be copied
            for sheet in sheets:
                sheet = book.worksheets[sheet]

                # Find the rows we need (Should have a date or time in the first column)
                rows_to_iterate = []
                i = 1
                for row in sheet.iter_rows(min_row=1, values_only=True):
                    cell_value = row[0]  # Access the first column of each row

                    if cell_value and is_date_or_time(str(cell_value)):
                        # The cell value is a date or time
                        rows_to_iterate.append(i)
                        i+=1
                    else:
                        # The cell value is not a date or time
                        i+=1

                # Find the date to use in the datetime
                row_num = 1; actual_date = 0
                for row in sheet.iter_rows(min_row=1, max_row = rows_to_iterate[0]-1, values_only=True):
                    if actual_date == 0:
                        col_num = 1
                        for cell in row:
                            cell_value = cell
                            if cell_value and is_date_or_time(str(cell_value)):
                                # The cell value is a date or time
                                row_with_date = row_num; col_with_date = col_num; actual_date = cell_value
                                col_num+=1
                                break
                            else:
                                # The cell value is not a date or time
                                # print(f"Not a date or time: {cell_value}")
                                col_num+=1
                        row_num += 1 

                # Move data into the main workbook
                for row in range(rows_to_iterate[0], rows_to_iterate[-1]+1):
                    for column in range(1, 27+1): # Columns A-AA
                        if column == 1: # Need to adjust the datetime to show the correct date
                            date = extract_date(str(actual_date))
                            cell_value = sheet.cell(row, column = 1).value
                            time = extract_time(str(cell_value))

                            if date and time:
                                combined_datetime = datetime.combine(date, time)
                                # print(f"Combined datetime: {combined_datetime}")
                                worksheet.cell(row = row_count+1, column = column).value = combined_datetime
                            else:
                                # print("Unable to extract both date and time components")
                                pass

                        else:
                            c = sheet.cell(row = row, column = column)
                            worksheet.cell(row = row_count+1, column = column).value = c.value
                            column += 1
                    row += 1; row_count += 1 
            
            os.remove(file)
            
            if stop == 0:
                stop_program()
        
    workbook.save(filename = main_file)
    text = Label(app, text = "Finished merging \n Saving, one moment")
    text.grid()
    stop_program()


def finishing_touches():
    global main_file_check; global main_data; global main_sheet; global workbook; global worksheet; global row_count; global column_count
    # Setting up the main file everything will go into
    if main_file_check == 0:
        main_data = pyxl.load_workbook(main_file)
        main_sheet = main_data.worksheets[0]

        # Setting up the active workbook
        global workbook
        workbook = Workbook()
        worksheet = workbook.active
        row_count = main_sheet.max_row
        column_count = main_sheet.max_column

    # Copy the heat number
    for row in range(5, row_count+1):
        c =  worksheet.cell(row-1,   2)
        c2 = worksheet.cell(row, 2)
        if c2.value == None:
            c2.value = c.value

    # Delete Empty Rows
    row_count_max = worksheet.max_row # Recount the rows
    rows_to_delete = []
    for row in range(4, row_count_max+1): # We must first find each row that needs to be deleted
        empty = []
        for column in range(4, 22+1): # columns D-V
            c = worksheet.cell(row, column)
            if c.value == None:
                empty.append(column)
        if len(empty) == 19: # 19 is all rows in range
            rows_to_delete.append(row)
    for row in reversed(rows_to_delete): # Delete the rows >:)
        worksheet.delete_rows(row)

    # Delete Rows that do not have heat numbers (Generally these are just rows that didn't need to be copied)
    row_count_max = worksheet.max_row # Recount the rows after rows have been deleted
    to_delete = []
    for row in range(4, row_count_max+1):
        c = worksheet.cell(row, 2) #Check the 2nd row for a heat number
        try:
            if int(c.value) not in range(100000, 9999999): #Range between 100,000 and 10,000,000 - just in case we go over 1,000,000 in heat numbers in the future
                to_delete.append(row)
        except:
            pass

    for row in reversed(to_delete):
        worksheet.delete_rows(row)


    ### Formatting
    # Merge cells
    worksheet.merge_cells('A1:A3'); worksheet.merge_cells('B1:B3'); worksheet.merge_cells('C1:C3')
    worksheet.merge_cells('D1:G1'); worksheet.merge_cells('D2:E2'); worksheet.merge_cells('F2:G2')
    worksheet.merge_cells('H1:S1'); worksheet.merge_cells('H2:M2'); worksheet.merge_cells('N2:P2'); worksheet.merge_cells('Q2:S2')
    worksheet.merge_cells('T1:T3'); worksheet.merge_cells('U1:U3'); worksheet.merge_cells('V1:V3'); worksheet.merge_cells('W1:Y3'); worksheet.merge_cells('X1:X3'); worksheet.merge_cells('Y1:Y3'); worksheet.merge_cells('Z1:Z3'); worksheet.merge_cells('AA1:AA3')
    for i in range(4, row_count+1):
        worksheet.merge_cells('W{}:Y{}'.format(i,i))

    # Center every cell
    for row in worksheet:
        for cell in row:
            cell.alignment = Alignment(horizontal = 'center', vertical = 'center')

    # Bold the font on the Header
    for row in range(1,3+1):
        for col in range(1, column_count+1):
            worksheet.cell(row = row, column = col).font = Font(bold=True)

    ### Apply borders
    # Put a thin border on every cell in the worksheet
    thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
    for i in range(1,row_count+1):
        for j in range(1, column_count):
            worksheet.cell(row = i, column = j).border = thin_border

    # Put a thick right border of select columns
    columns_for_border = [3, 7, 13, 19, 27]
    thick_border = Border(left=Side(style='thin'), 
                            right=Side(style='thick'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
    for col in columns_for_border:
        for i in range(1, row_count+1):
            worksheet.cell(row = i, column = col).border = thick_border

    # Add some special header borders
    columns_for_border = [2, 21, 22, 23, 24, 25, 26]
    bottom_header_border = Border(left=Side(style='thin'), 
                                    right=Side(style='thin'), 
                                    top=Side(style='thick'), 
                                    bottom=Side(style='thick'))
    for col in columns_for_border:
        worksheet.cell(row = 3, column = col).border = bottom_header_border

    columns_for_border = [5, 6, 9, 10, 11, 12, 15, 16, 17, 18]
    mid_header_border = Border(left=Side(style='thin'), 
                                right=Side(style='thin'), 
                                top=Side(style='thin'), 
                                bottom=Side(style='thick'))
    for col in columns_for_border:
        worksheet.cell(row = 3, column = col).border = mid_header_border

    columns_for_border = [1, 20]
    left_header_border = Border(left=Side(style='thick'), 
                                right=Side(style='thin'), 
                                top=Side(style='thick'), 
                                bottom=Side(style='thick'))
    for col in columns_for_border:
        worksheet.cell(row = 3, column = col).border = left_header_border

    columns_for_border = [3, 27]
    right_header_border = Border(left=Side(style='thin'), 
                                    right=Side(style='thick'), 
                                    top=Side(style='thick'), 
                                    bottom=Side(style='thick'))
    for col in columns_for_border:
        worksheet.cell(row = 3, column = col).border = right_header_border

    columns_for_border = [4, 8, 14]
    mid_left_header_border = Border(left=Side(style='thick'), 
                                right=Side(style='thin'), 
                                top=Side(style='thin'), 
                                bottom=Side(style='thick'))
    for col in columns_for_border:
        worksheet.cell(row = 3, column = col).border = mid_left_header_border

    columns_for_border = [7, 13, 19]
    mid_right_header_border = Border(left=Side(style='thin'), 
                                right=Side(style='thick'), 
                                top=Side(style='thin'), 
                                bottom=Side(style='thick'))
    for col in columns_for_border:
        worksheet.cell(row = 3, column = col).border = mid_right_header_border

def is_date_or_time(value):
    # Regex patterns for matching time and date/time formats
    time_pattern = r'^\d{2}:\d{2}:\d{2}$'
    datetime_pattern = r'^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}$'

    # Check if the value matches either time or datetime pattern
    if re.match(time_pattern, value) or re.match(datetime_pattern, value):
        return True
    else:
        return False

def extract_time(datetime_str):
    try:
        # Parse the datetime string
        datetime_value = datetime.strptime(datetime_str, "%Y-%m-%d %H:%M:%S")

        # Extract the time component
        time_component = datetime_value.time()

        return time_component
    except ValueError:
        # If the datetime string is in the format "HH:MM:SS"
        try:
            time_component = datetime.strptime(datetime_str, "%H:%M:%S").time()
            return time_component
        except ValueError:
            # Invalid datetime format
            return None

def extract_date(datetime_str):
    try:
        # Parse the datetime string
        datetime_value = datetime.strptime(datetime_str, "%Y-%m-%d %H:%M:%S")

        # Extract the date component
        date_component = datetime_value.date()

        return date_component
    except ValueError:
        # Invalid datetime format
        return None

# Functions for the addtional thread
def start_thread():
    # For repetition
    global main_file_check
    main_file_check = 0

    # Assign global variable and initialize value
    global stop
    stop = 1
    text = Label(app, text = 'Merging')
    text.grid()

    # Create and launch a thread 
    t = Thread (target = Merge_RPT_files)
    t.start()

def stop():
    # Assign global variable and set value to stop
    text = Label(app, text = 'Please wait, do not close the program yet')
    text.grid()
    global stop
    stop = 0

def stop_program():
    # Put on the finishing touches
    text = Label(app, text = 'Applying Formatting, this may take a few minutes')
    text.grid()
    finishing_touches()
    # Save the excel file and close the workbooks
    try:
        workbook.save(filename = main_file)
    except:
        pass
    try: # Just in case the backup is open for any reason
        os.chdir(dir_with_files)
        # shutil.copy(main_file, 'RPT Backup.xlsx')
    except:
        pass

    text = Label(app, text = "The program may be closed now")
    text.grid()
    sys.exit()

root_win = tk.Tk()
root_win.title("RPT Merge")
root_win.geometry('400x300')
app = Frame(root_win)
app.grid()
start_button = Button(app, text="Start the Merge",command=start_thread)
stop_button = Button(app, text="Stop Merging",command=stop)

start_button.grid()
stop_button.grid()

app.mainloop()
