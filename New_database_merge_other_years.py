# RPT Merge
# Note that this was created in Python and made into a windows executable file

import os, sys
from sys import exit
import fnmatch
import shutil
import pandas as pd
from threading import Thread
import openpyxl as pyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import *
from tkinter import simpledialog
from tkinter.filedialog import askopenfilename
import re
import time as timetime
from datetime import datetime, time

# Ask the User to give the files - this will allow this to work on any Windows computer
main_file = askopenfilename(title = 'File you want to update') # Excel File we want to add all the data into
dir_with_files = simpledialog.askstring(title = 'Folder path prompt', prompt = "Please type in the full path of the folder containing your files:    ")

def Merge_RPT_files():
    global main_file_check; global main_file; global main_data; global main_sheet; global workbook; global worksheet; global row_count; global column_count
    extra_save = 0 # For saving a backup automatically 

    # Setting up the main file everything will go into
    if main_file_check == 0:
        main_data = pyxl.load_workbook(main_file)
        main_sheet = main_data.worksheets[0]

        # Setting up the active workbook
        global workbook
        workbook = Workbook()
        worksheet = workbook.active
        row_count = main_sheet.max_row
        column_count = main_sheet.max_column
        if stop == 0:
            stop_program()

        # Make sure data in the main file stays (because this will overwite existing file)
        names = [sheetname.title for sheetname in main_data.worksheets]
        for name in names:
            main_sheet = main_data[name]
            workbook.create_sheet(name); worksheet = workbook[name]
            row_count = main_sheet.max_row; column_count = main_sheet.max_column
            for i in range (1, row_count + 1):
                for j in range (1, column_count + 1):
                    # reading cell value from source file
                    c = main_sheet.cell(row = i, column = j)
                    # writing the read value to destination file
                    worksheet.cell(row = i, column = j).value = c.value
                    # worksheet.cell(row = i, column = j).fill  = c.fill # keep any background filling
                    main_file_check = 1

    if stop == 0:
        stop_program()

    for folder in os.listdir(dir_with_files):
        folder = os.path.join(dir_with_files, folder)
        try:
            os.chdir(folder)
        except:
            continue
        for file in os.listdir(folder):
            try:
                book = pyxl.load_workbook(file, data_only = True)
            except:
                continue
            # names = book.sheetnames ### Does not work since there are charts in some workbooks (Counts as a sheetname but not a worksheet)
            names = [sheet.title for sheet in book.worksheets]
            # Have to get index for specific sheet
            i = 0
            source_sheets = []
            source_sheets = [i for i, name in enumerate(names) if name in [
                'INSP RPT', 'INSP RPT 1', 'INSP RPT 2', 'INSP RPT 3', 'INSP RPT 4', 'INSP RPT 5'
                'INSP RPT 1', 'INSP RPT 1 (2)', 'INSP RPT 2', 'INSP RPT 2 (2)',
                'INSP RPT 1 (8hr)', 'INSP RPT 1 (8hr) (2)', 'INSP RPT 1 (12hr)', 'INSP RPT 1 (12hr) (2)',
                'INSP RPT 2 (8hr)', 'INSP RPT 2 (8hr) (2)', 'INSP RPT 2 (12hr)', 'INSP RPT 2 (12hr) (2)',
                'INSP RPT 1-8', 'INSP RPT 1-8 (2)', 'INSP RPT 1-12', 'INSP RPT 1-12 (2)',
                'INSP RPT 2-8', 'INSP RPT 2-8 (2)', 'INSP RPT 2-12', 'INSP RPT 2-12 (2)'
            ]]

            # All sheets that need to be copied
            for sheet in source_sheets:
                sheet = book.worksheets[sheet]

                # We're going to grab the header from the source file so we can find the proper destination sheet
                header_end = 0
                for row in sheet.iter_rows(max_row = 20):
                    cell_value = row[0].value
                    if cell_value == 'Time' or cell_value == 'time' or cell_value == 'Hour' or cell_value == 'hour':
                        for cell in sheet.iter_cols(min_row = row[0].row, max_row = row[0].row):
                            cell_value = cell[0].value
                            if cell_value is not None and 'North Drift' in str(cell_value):
                                # print('Last Drift EQN is in row {} col {}'.format(cell[0].row, cell[0].column))
                                header_end = cell[0]
                                break

                # Earlier years may not have had a column for North Drift, and thus ended with comments
                if header_end == 0:
                    for row in sheet.iter_rows(max_row = 20):
                        cell_value = row[0].value
                        if cell_value == 'Time' or cell_value == 'time' or cell_value == 'Hour' or cell_value == 'hour':
                            for cell in sheet.iter_cols(min_row = row[0].row, max_row = row[0].row):
                                cell_value = cell[0].value
                                if cell_value is not None and ('comment' in str(cell_value).lower() or 'comments' in str(cell_value).lower()):
                                    # print('Last Drift EQN is in row {} col {}'.format(cell[0].row, cell[0].column))
                                    header_end = cell[0]
                                    break

                try:
                    # header_src = [[cell[0].value, cell[1].value, cell[2].value] for cell in sheet.iter_cols(min_row = header_end.row, max_row = header_end.row + 2, max_col = header_end.column)]
                    header_src = [[cell[0].value, cell[1].value] for cell in sheet.iter_cols(min_row = header_end.row, max_row = header_end.row + 2, max_col = header_end.column)]
                except:
                    continue

                # Now we need to get the header from all sheets in the destination file 
                sheets_dest = [sheet.title for sheet in workbook.worksheets]; sheets_dest_len = len(sheets_dest)
                header_same = 0 # We will use this later to see if any headers match
                header_end =  0
                for name in sheets_dest.copy():
                    # header_end = 0
                    worksheet = workbook[name]
                    for row in worksheet.iter_rows(max_row = 3):
                        cell_value = row[0].value
                        if cell_value == 'Time' or cell_value == 'time' or cell_value == 'Hour' or cell_value == 'hour':
                            for cell in worksheet.iter_cols(min_row = row[0].row, max_row = row[0].row):
                                cell_value = cell[0].value
                                if cell_value is not None and 'North Drift' in str(cell_value):
                                    # print('Last Drift EQN is in row {} col {}'.format(cell[0].row, cell[0].column))
                                    # header_end = cell[0] # For 2006, comment out this line
                                    break

                    if header_end == 0:
                        for name in sheets_dest.copy():
                            worksheet = workbook[name]
                            for row in worksheet.iter_rows(max_row = 3):
                                cell_value = row[0].value
                                if cell_value == 'Time' or cell_value == 'time' or cell_value == 'Hour' or cell_value == 'hour':
                                    for cell in worksheet.iter_cols(min_row = row[0].row, max_row = row[0].row):
                                        cell_value = cell[0].value
                                        if cell_value is not None and ('comment' in str(cell_value).lower() or 'comments' in str(cell_value).lower()):
                                            header_end = cell[0]
                                            break

                    # header_dest = [[cell[0].value, cell[1].value, cell[2].value] for cell in worksheet.iter_cols(min_row = header_end.row, max_row = header_end.row + 2, max_col = header_end.column)]
                    header_dest = [[cell[0].value, cell[1].value] for cell in worksheet.iter_cols(min_row = header_end.row, max_row = header_end.row + 2, max_col = header_end.column)]

                    if header_dest == header_src: # The headers are a match 
                        header_same = 1
                        worksheet = workbook[name]
                        break
                    else: # We need to check every name before doing making a new sheet
                        continue


                if header_same == 0: # no match for the headers
                    new_name = 'sheet {}'.format(sheets_dest_len+1)
                    workbook.create_sheet(new_name); names.append(new_name); sheets_dest_len+=1
                    worksheet = workbook[new_name]; j = 1
                    for col_val in header_src:
                        i = 1
                        for row_val in col_val:
                            worksheet.cell(row = i, column = j).value = row_val
                            i += 1
                        j += 1
                
                # Recount the rows and columns in case we switched sheets
                row_count = worksheet.max_row; column_count = worksheet.max_column

                # Find the rows we need (Should have a date or time in the first column)
                rows_to_iterate = []
                i = 1
                for row in sheet.iter_rows(min_row=1, values_only=True):
                    cell_value = row[0]  # Access the first column of each row

                    if cell_value and is_date_or_time(str(cell_value)):
                        # The cell value is a date or time
                        rows_to_iterate.append(i)
                        i+=1
                    elif str(cell_value).isnumeric():
                        rows_to_iterate.append(i)
                        i+=1
                    else:
                        # The cell value is not a date or time
                        i+=1

                # Find the date to use in the datetime
                row_num = 1; actual_date = 0
                try:
                    for row in sheet.iter_rows(min_row=1, max_row = rows_to_iterate[0]-1, values_only=True):
                        if actual_date == 0:
                            col_num = 1
                            for cell in row:
                                cell_value = cell
                                if cell_value and is_date_or_time(str(cell_value)):
                                    # The cell value is a date or time
                                    row_with_date = row_num; col_with_date = col_num; actual_date = cell_value
                                    col_num+=1
                                    break
                                else:
                                    # The cell value is not a date or time
                                    # print(f"Not a date or time: {cell_value}")
                                    col_num+=1
                            row_num += 1 
                except:
                    continue

                # Move data into the main workbook
                for row in range(rows_to_iterate[0], rows_to_iterate[-1]+1):
                    for column in range(1, 27+1): # Columns A-AA
                        if column == 1: # Need to adjust the datetime to show the correct date
                            date = extract_date(str(actual_date))
                            cell_value = sheet.cell(row, column = 1).value
                            time = extract_time(str(cell_value))

                            if date and time:
                                combined_datetime = datetime.combine(date, time)
                                # print(f"Combined datetime: {combined_datetime}")
                                worksheet.cell(row = row_count+1, column = column).value = combined_datetime
                            else:
                                # print("Unable to extract both date and time components")
                                combined_datetime = str(str(date) + ' hour ' + str(cell_value)) # Mostly just for 2006
                                worksheet.cell(row = row_count+1, column = column).value = combined_datetime
                                pass

                        else:
                            c = sheet.cell(row = row, column = column)
                            worksheet.cell(row = row_count+1, column = column).value = c.value
                            column += 1
                    row += 1; row_count += 1 
            
            os.remove(file)
            
            if stop == 0:
                stop_program()
        
    workbook.save(filename = main_file)
    text = Label(app, text = "Finished merging \n Saving, one moment")
    text.grid()
    stop_program()


def finishing_touches():
    global main_file_check; global main_file; global main_data; global main_sheet; global workbook; global worksheet; global row_count; global column_count
    # Setting up the main file everything will go into
    if main_file_check == 0:
        main_data = pyxl.load_workbook(main_file)
        main_sheet = main_data.worksheets[0]

        # Setting up the active workbook
        global workbook
        workbook = Workbook()
        worksheet = workbook.active

    # For each sheet
    names = [sheet.title for sheet in workbook.worksheets]
    names_len = len(names)
    for name in names.copy():
        worksheet = workbook[name]
        row_count = worksheet.max_row
        column_count = worksheet.max_column

        # Find the heat number
        for row in worksheet.iter_rows(max_row = 3):
            for cell in row:
                cell_value = cell.value
                if cell_value == 'Heat#' or cell_value == 'Heat #' or cell_value == 'Heat':
                    heat_col = cell.column
                    break
                else:
                    heat_col = 2

        # Copy the heat number 
        for row in range(5, row_count+1):
            c =  worksheet.cell(row-1, heat_col)
            c2 = worksheet.cell(row, heat_col)
            if c2.value == None:
                c2.value = c.value

    workbook.save(filename = main_file)

    # Delete the empty cells - using pandas becuase it is significantly faster
    df = pd.read_excel(main_file, header=None, sheet_name=None)

    directory = os.path.dirname(main_file)
    new_filename = os.path.splitext(main_file)[0]; new_filename = new_filename + '_final.xlsx'
    new_file = os.path.join(directory, new_filename) # Pandas has to save the file as a different name or else this does not work >:(
    writer = pd.ExcelWriter(new_file, engine='openpyxl') 

    for key in df:
        df_subset = df[key].iloc[3:]
        non_empty_rows = df_subset.dropna(how="all").dropna(subset=df[key].columns[3:22], how='all', inplace=False, thresh = 5)
        result = pd.concat([df[key].iloc[:3], non_empty_rows])
        result.to_excel(writer, key,index=False, header=False)
        
    writer.save()

    # set the main file to the new file to change minimal lines
    main_file = new_file

    # Reload the workbook
    main_data = pyxl.load_workbook(main_file)
    main_sheet = main_data.worksheets[0]
    # workbook = Workbook()
    workbook = main_data
    worksheet = workbook.active
    row_count = main_sheet.max_row
    column_count = main_sheet.max_column

# # Make sure data in the main file stays (because this will overwite existing file)
#     for i in range (1, row_count + 1):
#         for j in range (1, column_count + 1):
#             # reading cell value from source file
#             c = main_sheet.cell(row = i, column = j)
#             # writing the read value to destination file
#             worksheet.cell(row = i, column = j).value = c.value

    # Need to get the sheetnames again for some reason.  It's still there, so reason is unsure
    names = [sheet.title for sheet in workbook.worksheets]
    names_len = len(names)
    # Delete Rows that do not have heat numbers (Generally these are just rows that didn't need to be copied)
    for name in names.copy():
        worksheet = workbook[name] 
        row_count_max = worksheet.max_row # Recount the rows after rows have been deleted
        to_delete = []
        for row in range(4, row_count_max+1):
            c = worksheet.cell(row, 2) #Check the 2nd row for a heat number
            try:
                if int(c.value) not in range(100000, 9999999): #Range between 100,000 and 10,000,000 - just in case we go over 1,000,000 in heat numbers in the future
                    to_delete.append(row)
            except:
                pass

        for row in reversed(to_delete):
            worksheet.delete_rows(row)

    workbook.save(filename = main_file)

    # Let's delete any empty sheets (max row of 3)
    names = [sheet.title for sheet in workbook.worksheets]
    names_len = len(names)
    for sheet_name in names:
        sheet = workbook[sheet_name]
        # Check if the sheet is empty
        if sheet.max_row <= 3:
            # Delete the empty sheet
            workbook.remove(sheet)

    # Save the modified workbook
    workbook.save(filename = main_file)

    # Since each sheet has a different header format, we can't do much formatting automagically :(


def is_date_or_time(value):
    # Regex patterns for matching time and date/time formats
    time_pattern = r'^\d{2}:\d{2}:\d{2}$'
    datetime_pattern = r'^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}$'

    # Check if the value matches either time or datetime pattern
    if re.match(time_pattern, value) or re.match(datetime_pattern, value):
        return True
    else:
        return False

def extract_time(datetime_str):
    try:
        # Parse the datetime string
        datetime_value = datetime.strptime(datetime_str, "%Y-%m-%d %H:%M:%S")

        # Extract the time component
        time_component = datetime_value.time()

        return time_component
    except ValueError:
        # If the datetime string is in the format "HH:MM:SS"
        try:
            time_component = datetime.strptime(datetime_str, "%H:%M:%S").time()
            return time_component
        except ValueError:
            # Invalid datetime format
            return None

def extract_date(datetime_str):
    try:
        # Parse the datetime string
        datetime_value = datetime.strptime(datetime_str, "%Y-%m-%d %H:%M:%S")

        # Extract the date component
        date_component = datetime_value.date()

        return date_component
    except ValueError:
        # Invalid datetime format
        return None

# Functions for the addtional thread
def start_thread():
    # For repetition
    global main_file_check
    main_file_check = 0

    # Assign global variable and initialize value
    global stop
    stop = 1
    text = Label(app, text = 'Merging')
    text.grid()

    # Create and launch a thread 
    t = Thread (target = Merge_RPT_files)
    t.start()

def stop():
    # Assign global variable and set value to stop
    text = Label(app, text = 'Please wait, do not close the program yet')
    text.grid()
    global stop
    stop = 0

def stop_program():
    # Put on the finishing touches
    text = Label(app, text = 'Applying Formatting, this may take a few minutes')
    text.grid()
    finishing_touches()
    # Save the excel file and close the workbooks
    try:
        workbook.save(filename = main_file)
    except:
        pass
    try: # Just in case the backup is open for any reason
        os.chdir(dir_with_files)
        # shutil.copy(main_file, 'RPT Backup.xlsx')
    except:
        pass

    text = Label(app, text = "The program may be closed now")
    text.grid()
    sys.exit()


def save_program():
    # A button to save at any time in the process
    text = Label(app, text = 'Saving, one moment')
    text.grid()
    timetime.sleep(1)
    try:
        workbook.save(filename = main_file + ' backup.xlsx')
        text = Label(app, text = 'Saved')
        text.grid()
    except:
        text = Label(app, text = 'Error while saving, please try again')
        text.grid()


root_win = tk.Tk()
root_win.title("RPT Merge")
root_win.geometry('400x300')
app = Frame(root_win)
app.grid()
start_button = Button(app, text="Merge Files",command=start_thread)
stop_button = Button(app, text="Stop Merging",command=stop)
save_button = Button(app, text="Manual Save",command=save_program)

start_button.grid()
stop_button.grid()
save_button.grid()
app.mainloop()
